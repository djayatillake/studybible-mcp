#!/usr/bin/env python3
"""Ingest Burnett (Rom 4:18 / Gen 15:5 deification chain) and Gage/Bradley
(John ↔ Revelation typology) into cross_references and theology_content.

Idempotent: deletes existing rows where type IN ('gage','burnett') in
cross_references, and theology_content rows where source_work LIKE 'Burnett:%',
before inserting fresh.

Sources:
  data/bradley_gage_john_rev.xlsx  — Bradley/Gage John↔Rev typology spreadsheet
  data/Burnett-JSPL5.2.pdf         — Burnett, "So Shall Your Seed Be" JSPL 5.2 (2015)
"""

from __future__ import annotations

import datetime
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "db" / "study_bible.db"
XLSX_PATH = ROOT / "data" / "bradley_gage_john_rev.xlsx"

# Font colors in the Bradley/Gage xlsx:
#   FF000000 (black)              -> Bradley found it
#   FF741B47, FFA64D79 (purples)  -> from Warren Gage
#   theme:1 (default)             -> Bradley (treat unset as Bradley's contribution)
#   FF38761D (green, ~9 cells)    -> uncertain, mark Bradley?
GAGE_RGBS = {"FF741B47", "FFA64D79"}
BRADLEY_RGBS = {"FF000000"}


def cell_author(cell) -> str:
    """Bradley | Gage | Bradley? — derived from font color."""
    if cell.value is None:
        return "Bradley"
    f = cell.font
    if f and f.color:
        c = f.color
        if c.type == "rgb" and c.rgb:
            rgb = (c.rgb if isinstance(c.rgb, str) else "").upper()
            if rgb in GAGE_RGBS:
                return "Gage"
            if rgb in BRADLEY_RGBS:
                return "Bradley"
            if rgb == "FF38761D":
                return "Bradley?"
    return "Bradley"


def fix_value(v):
    """Excel mis-parses 'N:MM' chapter:verse strings as time. Reverse that."""
    if v is None:
        return None
    if isinstance(v, datetime.time):
        return f"{v.hour}:{v.minute:02d}" if v.minute else f"{v.hour}"
    s = str(v).strip()
    return s or None


def parse_refs(s: str | None) -> list[tuple[str, str]]:
    """Parse '1:6', '1:6-7', '1:14, 16-17' etc. into [(chapter, verse_or_range), ...].

    A bare '6' after '1:14, ' is interpreted as still in chapter 1.
    A range '6-7' is expanded to ['6', '7']. Ranges beyond that are left
    as the original 'a-b' for later expansion.
    """
    if not s:
        return []
    out: list[tuple[str, str]] = []
    cur_chap: str | None = None
    for raw_part in s.split(","):
        p = raw_part.strip().lstrip(".").strip()
        if not p:
            continue
        m = re.match(r"^(\d+):(\d+(?:-\d+)?)$", p)
        if m:
            cur_chap = m.group(1)
            out.append((cur_chap, m.group(2)))
            continue
        m = re.match(r"^(\d+):(\d+)", p)
        if m:
            cur_chap = m.group(1)
            out.append((cur_chap, m.group(2)))
            continue
        if re.match(r"^\d+(?:-\d+)?$", p) and cur_chap:
            out.append((cur_chap, p))
            continue
        # Unparseable token — skip
    return out


def expand_range(verse_or_range: str) -> list[int]:
    if "-" in verse_or_range:
        a, b = verse_or_range.split("-", 1)
        try:
            return list(range(int(a), int(b) + 1))
        except ValueError:
            return []
    try:
        return [int(verse_or_range)]
    except ValueError:
        return []


def to_usfm(book: str, chapter: int, verse: int) -> str:
    return f"{book}.{chapter}.{verse}"


# =============================================================================
# Bradley / Gage: John ↔ Revelation typology
# =============================================================================
def collect_gage_pairs() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    pairs: list[dict] = []
    skipped_unparsed = 0

    for sheet_name, tier in [
        ("Parallel John 1 - Rev 1", "parallel"),
        ("Chiastic John 1 - Rev 22", "chiastic"),
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_seen = False
        for row in ws.iter_rows():
            cells = list(row[:5]) + [None] * max(0, 5 - len(row))
            a, b, c, d, e = cells[:5]
            a_v = fix_value(a.value if a else None)
            b_v = fix_value(b.value if b else None)
            c_v = fix_value(c.value if c else None)
            d_v = fix_value(d.value if d else None)
            e_v = fix_value(e.value if e else None)

            if not a_v and not c_v:
                continue
            # Skip Bradley's running-commentary rows ("Looser connections, just noting")
            if a_v and "Looser connections" in a_v:
                continue
            if (a_v == "John") and (c_v in ("Rev", "Revelation")):
                header_seen = True
                continue
            if not header_seen:
                continue

            john_refs = parse_refs(a_v)
            rev_refs = parse_refs(c_v)
            if not john_refs or not rev_refs:
                skipped_unparsed += 1
                continue

            author_a = cell_author(a) if a else "Bradley"
            author_c = cell_author(c) if c else "Bradley"
            # Authorship of the pair: if either side is Gage, credit Gage
            authors = []
            if "Gage" in (author_a, author_c):
                authors.append("Gage")
            if "Bradley" in (author_a, author_c) or "Bradley?" in (author_a, author_c):
                authors.append("Bradley")
            attribution = "/".join(authors) if authors else "Bradley"

            theme_a = b_v or ""
            theme_c = d_v or ""
            commentary = e_v or ""

            for jc_s, jvr in john_refs:
                for jv in expand_range(jvr):
                    for rc_s, rvr in rev_refs:
                        for rv in expand_range(rvr):
                            try:
                                jc = int(jc_s)
                                rc = int(rc_s)
                            except ValueError:
                                continue
                            pairs.append({
                                "source": to_usfm("Jhn", jc, jv),
                                "target": to_usfm("Rev", rc, rv),
                                "tier": tier,
                                "theme_a": theme_a,
                                "theme_c": theme_c,
                                "commentary": commentary,
                                "attribution": attribution,
                            })

    print(f"  parsed {len(pairs)} canonical pairs ({skipped_unparsed} unparseable rows skipped)")
    return pairs


def gage_relevance(tier: str) -> int:
    # parallel: tighter, higher rank ; chiastic: explicit "looser connections"
    return {"parallel": 3, "chiastic": 1}.get(tier, 1)


_LEGEND_NOTES = (
    "Black is something I found",
    "Purple is from Dr. Warren Gage",
    "Looser connections",
)


def gage_note(p: dict) -> str:
    parts = []
    if p["theme_a"] or p["theme_c"]:
        parts.append(f"{p['theme_a']} ⟷ {p['theme_c']}".strip(" ⟷"))
    c = p["commentary"]
    if c and not any(legend in c for legend in _LEGEND_NOTES):
        if len(c) > 220:
            c = c[:217] + "..."
        parts.append(c)
    parts.append(p["attribution"])
    return " | ".join(parts)


# =============================================================================
# Burnett: Rom 4:18 / Gen 15:5 deification chain
# =============================================================================
# Hand-curated from Burnett, "So Shall Your Seed Be: Paul's Use of Genesis 15:5
# in Romans 4:18 in Light of Early Jewish Deification Traditions", JSPL 5.2
# (2015) 211-236. Each pair is part of the explicit interpretation chain Burnett
# argues for in the article. Pairs already in CH/TSK are still inserted under
# type='burnett' so the model can see Burnett's specific argument when filtering
# by source.
BURNETT_CITATION = "Burnett, JSPL 5.2 (2015) 211-236"

BURNETT_PAIRS = [
    # Gen 15:5 ↔ astral-glorification chain (the qualitative reading)
    ("Gen.15.5",  "Dan.12.3",  "Eschatological star-glory; Burnett pp. 226-228 (Dan 12:3 'shine like stars' is the same astral-immortality tradition Paul reads in Gen 15:5)"),
    ("Gen.15.5",  "Dan.12.2",  "Resurrection-vindication; Burnett p. 226 (Dan 12:1-3 as the eschatological frame for Gen 15:5)"),
    ("Gen.15.5",  "Mat.13.43", "Righteous shine like the sun; Burnett p. 232 (eschatological reaping echoes Dan 12:3, traceable to Gen 15:5 promise)"),
    ("Gen.15.5",  "Php.2.15",  "Children of God shine as lights; Burnett pp. 232-233 (Paul's own redeployment fusing Deut 32:5 + Dan 12:3 + Gen 15:5)"),
    ("Gen.15.5",  "Dan.8.10",  "Inverted host-of-heaven (false claimants cast down); Burnett p. 225"),
    # Divine-council frame (Deut 32 / Deut 4 inheritance)
    ("Gen.15.5",  "Deu.32.8",  "Abraham's seed inherits the bene-elohim role over the nations; Burnett pp. 230-232"),
    ("Gen.15.5",  "Deu.32.9",  "YHWH's portion-Jacob; Burnett p. 230"),
    ("Gen.15.5",  "Deu.4.19",  "Astral host allotted to nations — the role Abraham's seed inherits; Burnett pp. 230-232"),
    # Davidic / Balaam star-king tradition linking to Rom 4:13
    ("Gen.15.5",  "Num.24.17", "Balaam's 'star out of Jacob' — patriarchal vision and Davidic star-king share the promise; Burnett pp. 218-220"),
    ("Gen.15.5",  "Num.24.4",  "Both narrate seed-promise as a vision; Burnett p. 219"),
    # Saints judging angels = Abrahamic destiny fulfilled
    ("Gen.15.5",  "1Co.6.2",   "Saints judging the world; Burnett p. 235 (Abrahamic-seed eschatological office)"),
    ("Gen.15.5",  "1Co.6.3",   "Saints judging angels; Burnett p. 235 (taking over the bene-elohim's allotted role)"),
    # Gen 22:17 / 26:4 (reiteration cluster) — same chain
    ("Gen.22.17", "Dan.12.3",  "Reiteration of star-seed promise → Dan 12:3 shining; Burnett pp. 220-222 (Sirach 44:21 paraphrases this verse)"),
    ("Gen.22.17", "Mat.13.43", "Reiteration of star-seed → eschatological shining; Burnett pp. 220, 232"),
    ("Gen.22.17", "Php.2.15",  "Reiteration → Paul's 'shine as lights'; Burnett pp. 220, 232-233"),
    ("Gen.26.4",  "Dan.12.3",  "Reiteration of star-seed promise → Dan 12:3; Burnett p. 222"),
    ("Gen.26.4",  "Mat.13.43", "Reiteration → eschatological shining; Burnett p. 222"),
    # Rom 4:13 — "heir of the world" as Davidic-Balaam star-king inheritance
    ("Rom.4.13",  "Num.24.17", "'Heir of the world' = Davidic star-king tradition (Num 24:17 + Ps 72:8); Burnett pp. 219-220"),
    ("Rom.4.13",  "Psa.72.8",  "'From sea to sea' inheritance (paralleled in Sir 44:21); Burnett pp. 218-220"),
    ("Rom.4.13",  "Zec.9.10",  "Same 'sea to sea' inheritance; Burnett pp. 218-220"),
    # Rom 4:17-18 — "many nations" as Deut 32 inheritance, not just numerical
    ("Rom.4.17",  "Deu.32.8",  "'Many nations' = Deut 32 nations-allotment context, not just numerical; Burnett pp. 230-232"),
    ("Rom.4.17",  "Gen.15.5",  "Paul's pairing of Gen 17:5 + Gen 15:5 as one promise; Burnett pp. 213-214"),
    ("Rom.4.18",  "Deu.32.8",  "'Father of many nations' ↔ apportionment of nations; Burnett pp. 230-232"),
    ("Rom.4.18",  "Dan.12.3",  "Promise terminates in eschatological star-glory; Burnett pp. 226-228"),
    ("Rom.4.18",  "Num.24.17", "Davidic-Balaam star-tradition; Burnett pp. 218-220"),
    # Phil 2:15 internal links (Paul fusing Deut 32:5 + Dan 12:3)
    ("Php.2.15",  "Dan.12.3",  "Paul fuses Deut 32:5 + Dan 12:3; Burnett pp. 232-233 (Fee endorses Deut 32 echo)"),
    ("Php.2.15",  "Deu.32.5",  "Paul cites Deut 32:5 directly; Burnett p. 232"),
    # Job 38:7 — morning stars (heavenly host language)
    ("Gen.15.5",  "Job.38.7",  "Morning stars / heavenly host; Burnett p. 224"),
    # Sirach paraphrase context — note that Sirach itself is not indexed here
    # (canon-only), but the Davidic equivalent Sir 47:11 echoes 2 Sam 7
    ("Gen.15.5",  "2Sa.7.12",  "Davidic seed-promise paralleled to Abrahamic; Burnett pp. 220-222 (Sir 47:11 explicitly couples them)"),
]


# =============================================================================
# Burnett theology_content blocks (for the pseudepigraphal chain)
# =============================================================================
BURNETT_THEOLOGY = [
    {
        "source_work": "Burnett: So Shall Your Seed Be (JSPL 5.2, 2015)",
        "source_author": "burnett",
        "source_type": "article",
        "chapter_or_episode": "Main thesis",
        "title": "Abraham's Star-Like Seed: The Qualitative Reading of Genesis 15:5",
        "content_summary": (
            "Burnett argues that Paul's citation of Gen 15:5 in Rom 4:18 ('so shall your seed be') "
            "carries a qualitative as well as quantitative force, gone 'unnoticed or neglected' by "
            "most modern commentators. In a Second Temple stream of interpretation, Abraham's "
            "descendants are promised not merely innumerable seed but seed that is exalted *as the "
            "stars themselves* — i.e. functionally inheriting the cosmic-political role given to "
            "the angelic patrons of the nations in Deut 32:8 and Deut 4:19."
        ),
        "content_detail": (
            "Sirach 44:21 paraphrases the Abrahamic promise as: 'God promised him with an oath to "
            "bless the nations through his seed, to make him numerous as the grains of dust, and "
            "exalt (ἀνυψῶσαι) his seed AS THE STARS, giving them an inheritance from sea to sea, "
            "and from the River to the ends of the earth.' This is the exegetical pivot: Sirach "
            "reads Gen 22:17 not as 'as numerous as the stars' but as 'exalted as the stars' — "
            "qualitative status, not just multitude. Sirach 47:11 makes the same move for David. "
            "Apocalypse of Abraham 20:3-5 renarrates Gen 15:5 as a heavenly ascent in which "
            "Abraham counts the stars from above them, 'so shall I place for your seed the nation "
            "of men set apart for me.' 1 Enoch 104:2-6: 'But now you shall shine like the lights "
            "of heaven... you are about to be making a great rejoicing like the angels of heaven.' "
            "2 Baruch 51:10 (after the resurrection of 50:1-4): the splendor of the righteous "
            "'will be glorified in changes... so that they may be made equal to the angels and "
            "equal to the stars.' Daniel 12:3 names the same destiny ('the wise shall shine as "
            "the brightness of the firmament... like the stars forever'); Matt 13:43 ('the "
            "righteous will shine like the sun'); Phil 2:15 ('shine as lights in the world'). "
            "Each of these is rooted in the same star-seed tradition Paul activates with Gen 15:5."
        ),
        "page_range": "211-236",
        "url": None,
        "verses": [
            ("Gen.15.5", "primary"), ("Gen.22.17", "primary"), ("Gen.26.4", "primary"),
            ("Rom.4.18", "primary"), ("Rom.4.17", "primary"), ("Rom.4.13", "primary"),
            ("Dan.12.3", "primary"), ("Dan.12.2", "primary"),
            ("Mat.13.43", "primary"), ("Php.2.15", "primary"),
            ("Deu.32.8", "supporting"), ("Deu.32.9", "supporting"), ("Deu.4.19", "supporting"),
            ("Num.24.17", "supporting"), ("Num.24.4", "supporting"),
            ("Job.38.7", "supporting"),
            ("1Co.6.2", "supporting"), ("1Co.6.3", "supporting"),
        ],
    },
    {
        "source_work": "Burnett: So Shall Your Seed Be (JSPL 5.2, 2015)",
        "source_author": "burnett",
        "source_type": "article",
        "chapter_or_episode": "Second Temple background",
        "title": "Pseudepigraphal Witnesses to the Star-Seed Tradition",
        "content_summary": (
            "The Second Temple texts that read Abraham's star-seed promise qualitatively, not "
            "just as a count of descendants. Outside canon but the immediate background to "
            "Paul's argument in Romans 4."
        ),
        "content_detail": (
            "Sirach 44:21 — 'exalt his seed AS the stars, giving them inheritance from sea to "
            "sea.' Sirach 47:11 — 'the Lord exalted David's horn forever, gave him a covenant of "
            "kings and a glorious throne in Israel.' Apocalypse of Abraham 20:3-5 — Abraham "
            "ascends above the stars and is told 'as the number of the stars and their power so "
            "shall I place for your seed the nation of men set apart for me.' 1 Enoch 104:2-6 — "
            "the righteous 'shall shine like the lights of heaven... you are about to be making "
            "a great rejoicing like the angels of heaven.' 2 Baruch 21:4-5; 48:8-10; 50:1-4; "
            "51:10 — at the resurrection 'their splendor will be glorified in changes... they "
            "will be made equal to the angels and equal to the stars.' These are the texts the "
            "Pauline argument presupposes; their absence from modern commentaries on Rom 4 is "
            "what produces the purely-numerical reading Burnett is correcting."
        ),
        "page_range": "220-229",
        "url": None,
        "verses": [
            ("Gen.15.5", "primary"), ("Gen.22.17", "primary"), ("Gen.26.4", "primary"),
            ("Rom.4.18", "primary"), ("Rom.4.13", "supporting"),
            ("Dan.12.3", "primary"), ("Dan.12.2", "supporting"),
        ],
    },
    {
        "source_work": "Burnett: So Shall Your Seed Be (JSPL 5.2, 2015)",
        "source_author": "burnett",
        "source_type": "article",
        "chapter_or_episode": "Divine-council frame",
        "title": "Deuteronomy 32:8 and the Inheritance of the Nations",
        "content_summary": (
            "The Deut 32:8 (DSS/LXX) tradition — that YHWH apportioned the nations to the "
            "bene-elohim (sons of God) at Babel — is the structural frame for Burnett's reading "
            "of the Abrahamic promise. The seed of Abraham inherits the cosmic-political role "
            "those angelic patrons currently occupy."
        ),
        "content_detail": (
            "Deut 32:8-9 (DSS/LXX): 'When the Most High apportioned the nations... he fixed the "
            "boundaries of the peoples according to the number of the SONS OF GOD; for YHWH's "
            "portion is his people, Jacob his allotted heritage.' Deut 4:19: the sun, moon, and "
            "'all the host of heaven' have been 'allotted' (חלק) to all the peoples under "
            "heaven. In Second Temple thought (Sirach 17:17; Jubilees 15:31; Daniel 10) those "
            "celestial bodies ARE the angelic princes of the nations. The Abrahamic promise that "
            "his seed will be 'as the stars' is therefore a promise of inheritance: the "
            "resurrected righteous (Dan 12:3; Matt 13:43; Phil 2:15) take over the office held "
            "by the now-disinherited heavenly host (1 Cor 6:2-3, 'do you not know that we shall "
            "judge angels?'). Romans 4:13 calls Abraham 'heir of the world' for exactly this "
            "reason — the Balaam-Davidic star-king tradition (Num 24:17 / Ps 72:8 / Zech 9:10) "
            "describes the same dominion 'from sea to sea.'"
        ),
        "page_range": "230-235",
        "url": None,
        "verses": [
            ("Deu.32.8", "primary"), ("Deu.32.9", "primary"), ("Deu.4.19", "primary"),
            ("Gen.15.5", "primary"), ("Rom.4.13", "primary"), ("Rom.4.18", "primary"),
            ("1Co.6.2", "primary"), ("1Co.6.3", "primary"),
            ("Num.24.17", "supporting"), ("Psa.72.8", "supporting"), ("Zec.9.10", "supporting"),
            ("Dan.12.3", "supporting"), ("Mat.13.43", "supporting"), ("Php.2.15", "supporting"),
        ],
    },
]


# =============================================================================
# Insertion
# =============================================================================
def insert_xrefs(conn: sqlite3.Connection, gage_pairs: list[dict]):
    cur = conn.cursor()
    print("\nClearing existing rows where type IN ('gage','burnett')...")
    cur.execute("DELETE FROM cross_references WHERE type IN ('gage','burnett')")
    print(f"  deleted {cur.rowcount} rows")

    # Bradley/Gage — bidirectional
    print(f"\nInserting Bradley/Gage typology pairs ({len(gage_pairs)} forward × 2 = {2*len(gage_pairs)} rows)...")
    inserted = 0
    skipped_dupe = 0
    seen: set[tuple[str, str]] = set()
    for p in gage_pairs:
        for src, tgt in ((p["source"], p["target"]), (p["target"], p["source"])):
            key = (src, tgt)
            if key in seen:
                skipped_dupe += 1
                continue
            seen.add(key)
            cur.execute(
                "INSERT INTO cross_references (source, target, type, note, relevance) VALUES (?,?,?,?,?)",
                (src, tgt, "gage", gage_note(p), gage_relevance(p["tier"])),
            )
            inserted += 1
    print(f"  inserted {inserted} gage rows ({skipped_dupe} duplicates collapsed)")

    # Burnett — bidirectional
    print(f"\nInserting Burnett pairs ({len(BURNETT_PAIRS)} forward × 2 = {2*len(BURNETT_PAIRS)} rows)...")
    seen.clear()
    inserted = 0
    skipped_dupe = 0
    for src, tgt, note in BURNETT_PAIRS:
        full_note = f"{note} [{BURNETT_CITATION}]"
        for s, t in ((src, tgt), (tgt, src)):
            key = (s, t)
            if key in seen:
                skipped_dupe += 1
                continue
            seen.add(key)
            cur.execute(
                "INSERT INTO cross_references (source, target, type, note, relevance) VALUES (?,?,?,?,?)",
                (s, t, "burnett", full_note, 5),
            )
            inserted += 1
    print(f"  inserted {inserted} burnett rows ({skipped_dupe} duplicates collapsed)")
    conn.commit()


def insert_theology(conn: sqlite3.Connection):
    cur = conn.cursor()
    print("\nClearing existing Burnett theology_content rows...")
    cur.execute("DELETE FROM theology_verse_index WHERE content_id IN (SELECT id FROM theology_content WHERE source_author='burnett')")
    cur.execute("DELETE FROM theology_theme_index WHERE content_id IN (SELECT id FROM theology_content WHERE source_author='burnett')")
    cur.execute("DELETE FROM theology_content WHERE source_author='burnett'")
    print(f"  deleted {cur.rowcount} content rows (cascaded verse/theme index)")

    print(f"\nInserting {len(BURNETT_THEOLOGY)} Burnett theology blocks...")
    for entry in BURNETT_THEOLOGY:
        cur.execute(
            """INSERT INTO theology_content
               (source_work, source_author, source_type, chapter_or_episode, title,
                content_summary, content_detail, page_range, url)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (
                entry["source_work"], entry["source_author"], entry["source_type"],
                entry["chapter_or_episode"], entry["title"],
                entry["content_summary"], entry["content_detail"],
                entry["page_range"], entry["url"],
            ),
        )
        content_id = cur.lastrowid
        for ref, relevance in entry["verses"]:
            m = re.match(r"^(\w+)\.(\d+)\.(\d+)$", ref)
            if not m:
                print(f"    skipping malformed ref: {ref}")
                continue
            book, ch, vs = m.group(1), int(m.group(2)), int(m.group(3))
            cur.execute(
                """INSERT OR IGNORE INTO theology_verse_index
                   (content_id, reference, book, chapter, verse, relevance)
                   VALUES (?,?,?,?,?,?)""",
                (content_id, ref, book, ch, vs, relevance),
            )
        # Tag with deut32_worldview theme (already exists in theology_themes)
        cur.execute(
            """INSERT OR IGNORE INTO theology_theme_index (theme_key, content_id, reference)
               VALUES (?,?,?)""",
            ("deut32_worldview", content_id, None),
        )
    conn.commit()
    print("  done")


def main():
    if not DB_PATH.exists():
        print(f"DB not found: {DB_PATH}")
        sys.exit(1)
    if not XLSX_PATH.exists():
        print(f"xlsx not found: {XLSX_PATH}")
        sys.exit(1)

    print("=" * 60)
    print("Ingesting Burnett + Gage/Bradley")
    print("=" * 60)

    print("\nParsing Bradley/Gage xlsx with author colors...")
    gage_pairs = collect_gage_pairs()

    conn = sqlite3.connect(str(DB_PATH))
    try:
        insert_xrefs(conn, gage_pairs)
        insert_theology(conn)
    finally:
        conn.close()

    print("\nDone.")


if __name__ == "__main__":
    main()
